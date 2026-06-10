# paper_search.py
# ============================================================
# 📚 Article Search Tool — Google Scholar + PubMed
# Two independent keyword groups → one Google Sheets file
# (Group1 tab + Group2 tab), cumulative append
#
# v2 改動：
#   1. 擴充關鍵字 + 加入 MeSH terms
#   2. MAX_RESULTS 100（原 50）
#   3. 加入 AI prescreening layer（Claude Haiku）
#   4. COLUMNS 加入 AI Score / AI Reason 欄位
# ============================================================

import os
import time
import json
import datetime
import requests
import pandas as pd


# ── STEP 1: 🔑 CONFIG ─────────────────────────────────────

# --- Search Group 1：感染症 + AI/ML ---
KEYWORD_GROUPS_1 = [
    {
        "keywords": [
            "infectious disease",
            "infection control",
            "infection prevention and control",
            "antimicrobial stewardship",
            "antibiotic stewardship",
            "hospital-acquired infection",
            "nosocomial infection",
            "HAI",
            "sepsis",
            "bacteremia",
            "bloodstream infection",
        ],
        "logic": "OR",
    },
    {
        "keywords": [
            "machine learning",
            "deep learning",
            "neural network",
            "artificial intelligence",
            "generative AI",
            "large language model",
            "LLM",
            "GPT",
            "ChatGPT",
            "foundation model",
            "natural language processing",
            "clinical decision support",
            "predictive model",
        ],
        "logic": "OR",
    },
]
SHEET_TAB_1 = "Group1"

# --- Search Group 2：CR-hvKP ---
KEYWORD_GROUPS_2 = [
    {
        "keywords": [
            "crhvkp",
            "CR-hvKp",
            "CR-HvKP",
            "carbapenem resistant hypervirulent klebsiella pneumoniae",
            "carbapenem-resistant hypervirulent klebsiella",
            "hypervirulent klebsiella pneumoniae",
            "hvKP",
            "hvKp",
            "hypervirulent klebsiella",
        ],
        "logic": "OR",
    },
    {
        "keywords": [
            "clinical outcome",
            "mortality",
            "treatment",
            "infection",
            "bacteremia",
            "liver abscess",
        ],
        "logic": "OR",
    },
]
SHEET_TAB_2 = "Group2"

# --- Group descriptions（給 AI prescreening 用）---
GROUP_DESCRIPTIONS = {
    "Group1": (
        "Infectious disease clinical research combined with AI or machine learning. "
        "Relevant topics: sepsis prediction, antimicrobial stewardship AI tools, "
        "LLM applications in infection management, HAI surveillance models."
    ),
    "Group2": (
        "Carbapenem-resistant hypervirulent Klebsiella pneumoniae (CR-hvKP or crhvkp). "
        "Relevant topics: clinical outcomes, treatment, mortality, virulence factors, "
        "liver abscess, bacteremia caused by hypervirulent Klebsiella."
    ),
}

OUTPUT_FILE_1 = "articles_group1.xlsx"
OUTPUT_FILE_2 = "articles_group2.xlsx"

MAX_RESULTS   = 100   # v2：從 50 提高到 100
MONTHS_BACK_1 = 1
MONTHS_BACK_2 = 12

PUBMED_MAX_RETRIES = 3
PUBMED_RETRY_DELAY = 5
SCHOLAR_MAX_ERRORS = 3

# v2：加入 AI Score / AI Reason 欄位
COLUMNS = [
    "Status", "Source", "Title", "Abstract", "Publication Year",
    "Journal/Source", "DOI", "PMID", "PubMed URL",
    "AI Score", "AI Reason",
    "My Comment", "Drive link",
]


# ── STEP 2: Build query string ────────────────────────────

def build_query(groups):
    parts = []
    for group in groups:
        keywords = [kw.strip() for kw in group["keywords"] if kw.strip()]
        logic = group.get("logic", "AND").upper()
        if not keywords:
            continue
        if len(keywords) == 1:
            parts.append(keywords[0])
        else:
            joined = f" {logic} ".join(keywords)
            parts.append(f"({joined})")
    return " AND ".join(parts)


def preview_query(groups, label=""):
    query = build_query(groups)
    print("=" * 60)
    print(f"📋 Keyword groups {label}:")
    for i, g in enumerate(groups, 1):
        kws   = ", ".join(g["keywords"])
        logic = g.get("logic", "AND").upper()
        print(f" Group {i} [{logic}]: {kws}")
    print(f"\n🔎 Built query:\n {query}")
    print("=" * 60)
    return query


# ── STEP 3: Search PubMed ─────────────────────────────────

def _requests_get_with_retry(url, params, timeout=15):
    for attempt in range(1, PUBMED_MAX_RETRIES + 1):
        try:
            resp = requests.get(url, params=params, timeout=timeout)
            resp.raise_for_status()
            return resp
        except requests.RequestException as e:
            if attempt < PUBMED_MAX_RETRIES:
                print(f" ⚠️ Request failed (attempt {attempt}/{PUBMED_MAX_RETRIES}): {e}")
                print(f"    Retrying in {PUBMED_RETRY_DELAY}s...")
                time.sleep(PUBMED_RETRY_DELAY)
            else:
                print(f" ❌ All {PUBMED_MAX_RETRIES} attempts failed: {e}")
                raise


def search_pubmed(query, max_results=100, months_back=1):
    print("\n🔍 Searching PubMed...")
    results = []

    end_date   = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=30 * months_back)

    search_params = {
        "db":       "pubmed",
        "term":     query,
        "retmax":   max_results,
        "mindate":  start_date.strftime("%Y/%m/%d"),
        "maxdate":  end_date.strftime("%Y/%m/%d"),
        "datetype": "pdat",
        "retmode":  "json",
    }

    # 加入 API key（若有設定）
    api_key = os.environ.get("PUBMED_API_KEY")
    if api_key:
        search_params["api_key"] = api_key

    try:
        resp = _requests_get_with_retry(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
            search_params,
        )
        ids = resp.json().get("esearchresult", {}).get("idlist", [])
    except Exception as e:
        print(f" ❌ PubMed search failed, skipping: {e}")
        return results

    if not ids:
        print(" ⚠️ No PubMed results found for this date range.")
        return results

    print(f" ✅ Found {len(ids)} PubMed IDs — fetching details...")

    import xml.etree.ElementTree as ET

    for i in range(0, len(ids), 20):
        batch = ids[i : i + 20]
        fetch_params = {
            "db":      "pubmed",
            "id":      ",".join(batch),
            "retmode": "xml",
            "rettype": "abstract",
        }
        if api_key:
            fetch_params["api_key"] = api_key

        try:
            r    = _requests_get_with_retry(
                "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
                fetch_params,
            )
            root = ET.fromstring(r.text)
        except Exception as e:
            print(f" ❌ Failed to fetch batch {i // 20 + 1}, skipping: {e}")
            continue

        for article in root.findall(".//PubmedArticle"):
            try:
                title    = article.findtext(".//ArticleTitle") or ""
                abstract = " ".join(
                    a.text or "" for a in article.findall(".//AbstractText")
                )
                year = article.findtext(".//PubDate/Year") or \
                       article.findtext(".//PubDate/MedlineDate", "")[:4]
                journal  = article.findtext(".//Journal/Title") or \
                           article.findtext(".//MedlineTA") or ""
                pmid     = article.findtext(".//PMID") or ""
                doi_elem = article.find(".//ArticleId[@IdType='doi']")
                doi      = doi_elem.text.strip() if doi_elem is not None else ""

                results.append({
                    "Status":           "unread",
                    "Source":           "PubMed",
                    "Title":            title.strip(),
                    "Abstract":         abstract.strip(),
                    "Publication Year": year.strip(),
                    "Journal/Source":   journal.strip(),
                    "DOI":              doi,
                    "PMID":             pmid.strip(),
                    "PubMed URL":       f"https://pubmed.ncbi.nlm.nih.gov/{pmid.strip()}/"
                                        if pmid else "",
                    "AI Score":         "",
                    "AI Reason":        "",
                    "My Comment":       "",
                    "Drive link":       "",
                })
            except Exception as e:
                print(f" ⚠️ Skipped one article: {e}")

        time.sleep(0.4)

    return results


# ── STEP 4: Search Google Scholar (local only) ────────────

def search_google_scholar(query, max_results=100, months_back=1):
    try:
        from scholarly import scholarly as _scholarly
    except ImportError:
        print(" ❌ scholarly not installed, skipping Google Scholar.")
        return []

    print("🔍 Searching Google Scholar...")
    results = []

    cutoff_year        = (datetime.datetime.now() - datetime.timedelta(days=30 * months_back)).year
    consecutive_errors = 0

    try:
        search_query = _scholarly.search_pubs(query)
        count = 0
        while count < max_results:
            try:
                pub      = next(search_query)
                bib      = pub.get("bib", {})
                year_str = str(bib.get("pub_year", ""))

                if year_str:
                    try:
                        if int(year_str) < cutoff_year:
                            continue
                    except ValueError:
                        pass

                doi = (
                    bib.get("doi", "")
                    or pub.get("externalIds", {}).get("DOI", "")
                    or ""
                )

                results.append({
                    "Status":           "unread",
                    "Source":           "Google Scholar",
                    "Title":            bib.get("title", "").strip(),
                    "Abstract":         bib.get("abstract", "").strip(),
                    "Publication Year": year_str.strip(),
                    "Journal/Source":   (
                        bib.get("venue", "") or bib.get("journal", "")
                    ).strip(),
                    "DOI":              doi.strip(),
                    "PMID":             "",
                    "PubMed URL":       "",
                    "AI Score":         "",
                    "AI Reason":        "",
                    "My Comment":       "",
                    "Drive link":       "",
                })
                count += 1
                consecutive_errors = 0
                time.sleep(1.2)

            except StopIteration:
                break
            except Exception as e:
                consecutive_errors += 1
                err_msg = str(e).lower()
                if any(kw in err_msg for kw in ["captcha", "blocked", "429", "forbidden", "robot"]):
                    print(f" ❌ Google Scholar blocked: {e}")
                    break
                print(f" ⚠️ Skipped one result ({consecutive_errors}/{SCHOLAR_MAX_ERRORS}): {e}")
                if consecutive_errors >= SCHOLAR_MAX_ERRORS:
                    print(f" ❌ Too many consecutive errors, stopping Scholar search.")
                    break
                time.sleep(2)

    except Exception as e:
        print(f" ❌ Google Scholar error: {e}")
        print(" ℹ️  Continuing with PubMed results only.")

    print(f" ✅ Retrieved {len(results)} results from Google Scholar.")
    return results


# ── STEP 5: AI Prescreening（Claude Haiku）─────────────────

def ai_prescreening(results, group_description, sheet_tab):
    """
    用 Claude Haiku 對每篇 title + abstract 打分 0-2。
    0 = 明顯不相關（排除）
    1 = 可能相關（保留，給人工看）
    2 = 直接相關（保留，優先看）

    需要環境變數 ANTHROPIC_API_KEY。
    若未設定則跳過，全部保留原始結果。
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print(" ℹ️  ANTHROPIC_API_KEY 未設定，跳過 AI prescreening，保留全部結果。")
        return results

    try:
        import anthropic
    except ImportError:
        print(" ⚠️ anthropic 套件未安裝，跳過 AI prescreening。")
        print("    執行：pip install anthropic")
        return results

    client  = anthropic.Anthropic(api_key=api_key)
    kept    = []
    removed = 0

    print(f"\n🤖 AI prescreening（Claude Haiku）— {len(results)} 篇...")

    for i, paper in enumerate(results, 1):
        title    = paper.get("Title", "")[:300]
        abstract = paper.get("Abstract", "")[:600]

        prompt = f"""You are an infectious disease physician screening research papers.

Research focus: {group_description}

Paper:
Title: {title}
Abstract: {abstract}

Score this paper:
2 = Directly relevant to the research focus — should read
1 = Possibly relevant or borderline — keep for manual review
0 = Clearly irrelevant — can be excluded

Reply with ONLY valid JSON, no extra text:
{{"score": 0, "reason": "one short sentence explaining the score"}}"""

        try:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=120,
                messages=[{"role": "user", "content": prompt}],
            )
            raw    = response.content[0].text.strip()
            result = json.loads(raw)
            score  = int(result.get("score", 1))
            reason = str(result.get("reason", ""))
        except Exception as e:
            # 解析失敗 → 保留（預設分數 1）
            score  = 1
            reason = f"parse error: {e}"

        paper["AI Score"]  = score
        paper["AI Reason"] = reason

        if score > 0:
            kept.append(paper)
        else:
            removed += 1

        # 進度顯示（每 10 篇一次）
        if i % 10 == 0 or i == len(results):
            print(f"   {i}/{len(results)} screened — kept {len(kept)}, removed {removed}")

        time.sleep(0.3)  # 避免 rate limit

    print(f" ✅ AI prescreening 完成：{len(results)} → {len(kept)} 篇（移除 {removed} 篇不相關）")
    return kept


# ── STEP 6: Save to Excel ─────────────────────────────────

def save_to_excel(df, output_file, query, months_back):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    df.to_excel(output_file, index=False, engine="openpyxl")
    wb = load_workbook(output_file)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = {
        "A": 12, "B": 15, "C": 55, "D": 80, "E": 12,
        "F": 35, "G": 35, "H": 14, "I": 42,
        "J": 10, "K": 40,   # AI Score, AI Reason
        "L": 22, "M": 42,   # My Comment, Drive link
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Abstract wrap
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # AI Reason wrap
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Status 色彩
    status_colors = {"keep": "C6EFCE", "skip": "FFCCCC", "unread": "FFFFFF"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            color = status_colors.get(str(cell.value).strip().lower(), "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # AI Score 色彩（J 欄）
    score_colors = {"2": "C6EFCE", "1": "FFEB9C", "0": "FFCCCC"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10):
        for cell in row:
            color = score_colors.get(str(cell.value).strip(), "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_note = wb.create_sheet("Search Info")
    ws_note["A1"], ws_note["B1"] = "Search Query",  query
    ws_note["A2"], ws_note["B2"] = "Date Range",    f"Last {months_back} month(s)"
    ws_note["A3"], ws_note["B3"] = "Run Date",      datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_note["A4"], ws_note["B4"] = "Total Articles", len(df)

    wb.save(output_file)
    print(f" ✅ Excel saved: {output_file}")


# ── STEP 7: Append to Google Sheets ──────────────────────

def append_to_sheets(df, sheet_tab, run_date=None):
    creds_json     = os.environ.get("GDRIVE_CREDENTIALS")
    spreadsheet_id = os.environ.get("GSHEET_SPREADSHEET_ID")

    if not creds_json or not spreadsheet_id:
        print("⚠️  GDRIVE_CREDENTIALS or GSHEET_SPREADSHEET_ID not set, skipping Sheets upload.")
        return

    try:
        from googleapiclient.discovery import build
        from google.oauth2 import service_account
    except ImportError:
        print("❌ google-api-python-client not installed.")
        return

    creds = service_account.Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    service = build("sheets", "v4", credentials=creds)
    sheet   = service.spreadsheets()

    if run_date is None:
        run_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # 讀取現有 title 做去重（C 欄 = Title）
    try:
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_tab}!C:C",
        ).execute()
        existing_values = result.get("values", [])
    except Exception:
        existing_values = []

    existing_titles = set(
        row[0].strip().lower()
        for row in existing_values[1:]
        if row
    )

    headers = ["Batch date"] + COLUMNS

    # 寫入表頭（若空白）
    try:
        first_cell = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_tab}!A1",
        ).execute()
        has_header = bool(first_cell.get("values"))
    except Exception:
        has_header = False

    if not has_header:
        sheet.values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_tab}!A1",
            valueInputOption="RAW",
            body={"values": [headers]},
        ).execute()
        print(f" ✅ Header written to tab '{sheet_tab}'")

    # 過濾重複，組成新列
    new_rows = []
    skipped  = 0
    for _, row in df.iterrows():
        title = str(row.get("Title", "")).strip().lower()
        if title in existing_titles:
            skipped += 1
            continue
        new_rows.append(
            [run_date] + [str(row.get(col, "")) for col in COLUMNS]
        )

    if skipped:
        print(f" ℹ️  Skipped {skipped} duplicate title(s)")

    if not new_rows:
        print(f" ℹ️  No new rows to append to tab '{sheet_tab}'")
        return

    sheet.values().append(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_tab}!A1",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": new_rows},
    ).execute()

    print(f" ✅ Appended {len(new_rows)} new row(s) to Sheets tab '{sheet_tab}'")


# ── STEP 8: Run a single search group ────────────────────

def run_search(
    keyword_groups,
    output_file,
    sheet_tab,
    label="",
    max_results=MAX_RESULTS,
    months_back=1,
):
    query    = preview_query(keyword_groups, label=label)
    run_date = datetime.datetime.now().strftime("%Y-%m-%d")

    print(f"\n 📅 Date range : Last {months_back} month(s)")
    print(f" 📦 Max results : {max_results} per source\n")

    # PubMed（永遠執行）
    pubmed_results = search_pubmed(query, max_results, months_back)

    # Google Scholar（本機才跑）
    skip_scholar = os.environ.get("SKIP_SCHOLAR", "false").lower() == "true"
    if skip_scholar:
        print("ℹ️  Google Scholar skipped (SKIP_SCHOLAR=true)")
        scholar_results = []
    else:
        scholar_results = search_google_scholar(query, max_results, months_back)

    all_results = pubmed_results + scholar_results

    if not all_results:
        print(f"\n❌ No articles found for {label}.")
        return None

    # ── v2：AI prescreening ──
    group_desc  = GROUP_DESCRIPTIONS.get(sheet_tab, label)
    all_results = ai_prescreening(all_results, group_desc, sheet_tab)

    if not all_results:
        print(f"\n❌ All articles filtered out by AI prescreening for {label}.")
        return None

    df = pd.DataFrame(all_results, columns=COLUMNS)

    before = len(df)
    df.drop_duplicates(subset="Title", keep="first", inplace=True)
    df.reset_index(drop=True, inplace=True)
    after = len(df)
    if before != after:
        print(f"\n🧹 Removed {before - after} duplicate(s) within batch.")

    # AI Score 降冪排序（讓分數高的排前面，方便人工篩選）
    if "AI Score" in df.columns:
        df["AI Score"] = pd.to_numeric(df["AI Score"], errors="coerce").fillna(1)
        df.sort_values("AI Score", ascending=False, inplace=True)
        df.reset_index(drop=True, inplace=True)

    save_to_excel(df, output_file, query, months_back)

    print(f"\n📤 Appending to Google Sheets tab '{sheet_tab}'...")
    append_to_sheets(df, sheet_tab, run_date)

    print(f"\n✅ Done! {after} articles processed for {label}")
    print(f" • PubMed        : {len(pubmed_results)} articles (before AI screen)")
    print(f" • Google Scholar: {len(scholar_results)} articles (before AI screen)")
    print(f" • After AI screen: {after} articles")

    return df


# ── STEP 9: Run both groups ───────────────────────────────

if __name__ == "__main__":
    print("\n" + "🟦" * 30)
    print("  Running Search Group 1")
    print("🟦" * 30)
    run_search(
        keyword_groups=KEYWORD_GROUPS_1,
        output_file=OUTPUT_FILE_1,
        sheet_tab=SHEET_TAB_1,
        label="(Group 1)",
        months_back=MONTHS_BACK_1,
    )

    print("\n" + "🟩" * 30)
    print("  Running Search Group 2")
    print("🟩" * 30)
    run_search(
        keyword_groups=KEYWORD_GROUPS_2,
        output_file=OUTPUT_FILE_2,
        sheet_tab=SHEET_TAB_2,
        label="(Group 2)",
        months_back=MONTHS_BACK_2,
    )

    print("\n🎉 All done!")
